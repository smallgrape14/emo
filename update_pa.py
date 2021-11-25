import os
import json
import re
import time
import math
import random
import requests
from openpyxl import load_workbook, Workbook
import random
import string
#-*- coding: UTF-8 -*-
from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配
import urllib.request, urllib.error  # 制定URL，获取网页数据
import xlwt  # 进行excel操作
import xlwt, xlrd
from xlutils.copy import copy
match=re.compile(r'<p(.*?)>(.*?)</p>')
id=0
# 数据采集-------------------------------------修改部分和url
table=["失望","焦虑"]
num=2
title="失望焦虑时鼓励"
#-----------------------------------------
def getinfo(url, headers):
    allinfo = []
    global  id
    global  match
    global num
    global table
    try:
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        r.encoding = 'utf-8'
        datas = json.loads(r.text)
        for info in datas['data']:
            text = info['content']  # 拿到一个网友的回答内容
            newer=re.findall(match,text)      #过滤
            for item  in newer:
                i=item[1]
                # if(i.isspace()==''):
                #     print("kongkong")
                #     continue
                i = i.strip(' */<>')
                i=i.strip(string.digits)  # 输出包含数字0~9的字符串
                i=i.strip(string.ascii_letters)  # 包含所有字母(大写或小写)的字符串
                i=i.strip('!"#$%&()*+,-./:;<=>?@[\\]^_‘{|}~、')
                index=random.randint(0,num-1)
                oneinfo=[str(id),str(i),table[index]]
                allinfo.append(oneinfo)
                id=id+1
                print(oneinfo)
        print("--------------------")
        # print(allinfo)
        next_url = datas['paging']['next']
        saveData(filepath, allinfo)
        # return
        if datas['paging']['is_end']:
            print('----')
            return

        time.sleep(random.uniform(5.1, 20.1))
        return getinfo(next_url, headers)
    except:
        return getinfo(next_url, headers)

#数据保存
def saveData( filename,datalist):
    print("save excel.......")
    # print(datalist)
    global num
    data = xlrd.open_workbook(filename, formatting_info=True)
    excel = copy(wb=data)  # 完成xlrd对象向xlwt对象转换
    excel_table = excel.get_sheet(0)  # 获得要操作的页
    table = data.sheets()[0]
    nrows = table.nrows  # 获得行数
    # ncols = table.ncols  # 获得列数
    # print(datalist)
    for value in datalist:
        # list=[str(1),value,'悲伤']
        print(value)
        for i in range(3):
            excel_table.write(nrows, i, value[i])  # 因为单元格从0开始算，所以row不需要加一
        # print(value)
        nrows = nrows + 1
    excel.save(filename)
# 数据保存
def insert2excel(filepath, allinfo):
    try:
        if not os.path.exists(filepath):
            print(2)
            tableTitle = ['pushmsg_id', 'content','tips']
            wb = Workbook()
            ws = wb.active
            ws.title = 'sheet1'
            ws.append(tableTitle)
            wb.save(filepath)
            time.sleep(3)
        wb = load_workbook(filepath)
        ws = wb.active
        ws.title = 'sheet1'
        for info in allinfo:
            ws.append(info)
        wb.save(filepath)
        print('文件已更新')
    except:
        print('文件更新失败')


# url = 'https://www.zhihu.com/api/v4/questions/368550554/answers?include=data%5B%2A%5D.is_normal%2Cadmin_closed_comment%2Creward_info%2Cis_collapsed%2Cannotation_action%2Cannotation_detail%2Ccollapse_reason%2Cis_sticky%2Ccollapsed_by%2Csuggest_edit%2Ccomment_count%2Ccan_comment%2Ccontent%2Ceditable_content%2Cattachment%2Cvoteup_count%2Creshipment_settings%2Ccomment_permission%2Ccreated_time%2Cupdated_time%2Creview_info%2Crelevant_info%2Cquestion%2Cexcerpt%2Cis_labeled%2Cpaid_info%2Cpaid_info_content%2Crelationship.is_authorized%2Cis_author%2Cvoting%2Cis_thanked%2Cis_nothelp%2Cis_recognized%3Bdata%5B%2A%5D.mark_infos%5B%2A%5D.url%3Bdata%5B%2A%5D.author.follower_count%2Cvip_info%2Cbadge%5B%2A%5D.topics%3Bdata%5B%2A%5D.settings.table_of_content.enabled&limit=20&offset=0&platform=desktop&sort_by=default'
#url='https://www.zhihu.com/api/v4/questions/442490819/answers?include=data%5B%2A%5D.is_normal%2Cadmin_closed_comment%2Creward_info%2Cis_collapsed%2Cannotation_action%2Cannotation_detail%2Ccollapse_reason%2Cis_sticky%2Ccollapsed_by%2Csuggest_edit%2Ccomment_count%2Ccan_comment%2Ccontent%2Ceditable_content%2Cattachment%2Cvoteup_count%2Creshipment_settings%2Ccomment_permission%2Ccreated_time%2Cupdated_time%2Creview_info%2Crelevant_info%2Cquestion%2Cexcerpt%2Cis_labeled%2Cpaid_info%2Cpaid_info_content%2Crelationship.is_authorized%2Cis_author%2Cvoting%2Cis_thanked%2Cis_nothelp%2Cis_recognized%3Bdata%5B%2A%5D.mark_infos%5B%2A%5D.url%3Bdata%5B%2A%5D.author.follower_count%2Cvip_info%2Cbadge%5B%2A%5D.topics%3Bdata%5B%2A%5D.settings.table_of_content.enabled&limit=5&offset=5&platform=desktop&sort_by=default'
# url="https://www.zhihu.com/api/v4/questions/399495868/answers?include=data%5B%2A%5D.is_normal,admin_closed_comment,reward_info,is_collapsed,annotation_action,annotation_detail,collapse_reason,is_sticky,collapsed_by,suggest_edit,comment_count,can_comment,content,editable_content,attachment,voteup_count,reshipment_settings,comment_permission,created_time,updated_time,review_info,relevant_info,question,excerpt,is_labeled,paid_info,paid_info_content,relationship.is_authorized,is_author,voting,is_thanked,is_nothelp,is_recognized;data[*].mark_infos[*].url;data[*].author.follower_count,vip_info,badge[*].topics;data[*].settings.table_of_content.enabled&limit=5&offset=5&platform=desktop&sort_by=default"

url="https://www.zhihu.com/api/v4/questions/450784111/answers?include=data[*].is_normal,admin_closed_comment,reward_info,is_collapsed,annotation_action,annotation_detail,collapse_reason,is_sticky,collapsed_by,suggest_edit,comment_count,can_comment,content,editable_content,attachment,voteup_count,reshipment_settings,comment_permission,created_time,updated_time,review_info,relevant_info,question,excerpt,is_labeled,paid_info,paid_info_content,relationship.is_authorized,is_author,voting,is_thanked,is_nothelp,is_recognized;data[*].mark_infos[*].url;data[*].author.follower_count,vip_info,badge[*].topics;data[*].settings.table_of_content.enabled&limit=5&offset=0&platform=desktop&sort_by=default"
url="https://www.zhihu.com/api/v4/questions/389815242/answers?include=data%5B%2A%5D.is_normal,admin_closed_comment,reward_info,is_collapsed,annotation_action,annotation_detail,collapse_reason,is_sticky,collapsed_by,suggest_edit,comment_count,can_comment,content,editable_content,attachment,voteup_count,reshipment_settings,comment_permission,created_time,updated_time,review_info,relevant_info,question,excerpt,is_labeled,paid_info,paid_info_content,relationship.is_authorized,is_author,voting,is_thanked,is_nothelp,is_recognized;data[*].mark_infos[*].url;data[*].author.follower_count,vip_info,badge[*].topics;data[*].settings.table_of_content.enabled&limit=5&offset=0&platform=desktop&sort_by=default"
url="https://www.zhihu.com/api/v4/questions/376530821/answers?include=data[*].is_normal,admin_closed_comment,reward_info,is_collapsed,annotation_action,annotation_detail,collapse_reason,is_sticky,collapsed_by,suggest_edit,comment_count,can_comment,content,editable_content,attachment,voteup_count,reshipment_settings,comment_permission,created_time,updated_time,review_info,relevant_info,question,excerpt,is_labeled,paid_info,paid_info_content,relationship.is_authorized,is_author,voting,is_thanked,is_nothelp,is_recognized;data[*].mark_infos[*].url;data[*].author.follower_count,vip_info,badge[*].topics;data[*].settings.table_of_content.enabled&limit=5&offset=0&platform=desktop&sort_by=default"
headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36',
}
filepath = "C://Users//duan//Desktop//pa_db//"+title+".xls"
workbook = xlwt.Workbook(encoding='utf-8') #创建workbook 对象
worksheet = workbook.add_sheet('sheet1') #创建工作表sheet
# worksheet.write(0, 0, 'hello') #往表中写内容,第一各参数 行,第二个参数列,第三个参数内容
workbook.save(filepath) #保存表为students.xls

print(filepath)
getinfo(url, headers)
# datalist=[[1,1,1],[2,3,4],[2,3,4]]
# saveData(filepath,datalist)
