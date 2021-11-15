import os
import json
import re
import time
import math
import random
import requests
from openpyxl import load_workbook, Workbook

match=re.compile(r'<p>(.*?)</p>')
id=0
# 数据采集
def getinfo(url, headers):
    allinfo = []
    global  id
    global  match
    try:
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        r.encoding = 'utf-8'
        datas = json.loads(r.text)
        for info in datas['data']:
            text = info['content']  # 拿到一个网友的回答内容
            newer=re.findall(match,text)      #过滤
            for item  in newer:
                 if '\u4e00' <= item <= '\u9fff':
                    oneinfo=[str(id),item]
                    allinfo.append(oneinfo)
                    id=id+1
                    print(oneinfo)
            return
        next_url = datas['paging']['next']
        insert2excel(filepath, allinfo)

        if datas['paging']['is_end']:
            print('----')
            return

        time.sleep(random.uniform(5.1, 20.1))
        return getinfo(next_url, headers)
    except:
        return getinfo(next_url, headers)


# 数据保存
def insert2excel(filepath, allinfo):
    try:
        if not os.path.exists(filepath):
            tableTitle = ['id', '内容']
            wb = Workbook()
            ws = wb.active
            ws.title = 'sheet1'
            ws.append(tableTitle)
            wb.save(filepath)
            time.sleep(3)
        wb = load_workbook(filepath)
        ws = wb.active
        ws.title = 'sheet1'
        for info in allinfo:
            ws.append(info)
        wb.save(filepath)
        print('文件已更新')
    except:
        print('文件更新失败')


# url = 'https://www.zhihu.com/api/v4/questions/368550554/answers?include=data%5B%2A%5D.is_normal%2Cadmin_closed_comment%2Creward_info%2Cis_collapsed%2Cannotation_action%2Cannotation_detail%2Ccollapse_reason%2Cis_sticky%2Ccollapsed_by%2Csuggest_edit%2Ccomment_count%2Ccan_comment%2Ccontent%2Ceditable_content%2Cattachment%2Cvoteup_count%2Creshipment_settings%2Ccomment_permission%2Ccreated_time%2Cupdated_time%2Creview_info%2Crelevant_info%2Cquestion%2Cexcerpt%2Cis_labeled%2Cpaid_info%2Cpaid_info_content%2Crelationship.is_authorized%2Cis_author%2Cvoting%2Cis_thanked%2Cis_nothelp%2Cis_recognized%3Bdata%5B%2A%5D.mark_infos%5B%2A%5D.url%3Bdata%5B%2A%5D.author.follower_count%2Cvip_info%2Cbadge%5B%2A%5D.topics%3Bdata%5B%2A%5D.settings.table_of_content.enabled&limit=20&offset=0&platform=desktop&sort_by=default'
url='https://www.zhihu.com/api/v4/questions/442490819/answers?include=data%5B%2A%5D.is_normal%2Cadmin_closed_comment%2Creward_info%2Cis_collapsed%2Cannotation_action%2Cannotation_detail%2Ccollapse_reason%2Cis_sticky%2Ccollapsed_by%2Csuggest_edit%2Ccomment_count%2Ccan_comment%2Ccontent%2Ceditable_content%2Cattachment%2Cvoteup_count%2Creshipment_settings%2Ccomment_permission%2Ccreated_time%2Cupdated_time%2Creview_info%2Crelevant_info%2Cquestion%2Cexcerpt%2Cis_labeled%2Cpaid_info%2Cpaid_info_content%2Crelationship.is_authorized%2Cis_author%2Cvoting%2Cis_thanked%2Cis_nothelp%2Cis_recognized%3Bdata%5B%2A%5D.mark_infos%5B%2A%5D.url%3Bdata%5B%2A%5D.author.follower_count%2Cvip_info%2Cbadge%5B%2A%5D.topics%3Bdata%5B%2A%5D.settings.table_of_content.enabled&limit=5&offset=5&platform=desktop&sort_by=default'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36',
}
filepath = 'emo_send.xlsx'
getinfo(url, headers)
